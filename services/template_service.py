"""Generate long-format Excel templates per dataset (Fase 1)."""

from __future__ import annotations

import re
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from services.dataset_catalog import DatasetDefinition, dataset_workbook_sheet_title, get_dataset, iter_datasets

if TYPE_CHECKING:
    from flask import Response

HEADER_FONT = Font(name="Arial", size=11, bold=True)
BODY_FONT = Font(name="Arial", size=11, color="000000")


def _write_long_sheet(wb: Workbook, definition: DatasetDefinition, *, with_sample: bool) -> None:
    title = dataset_workbook_sheet_title(definition)
    ws = wb.create_sheet(title)
    headers = definition.required_template_headers
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    rows = definition.sample_rows if with_sample else ()
    for r_idx, row in enumerate(rows, start=2):
        if len(row) != len(headers):
            raise ValueError(f"sample_rows width mismatch for {definition.slug}")
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        mx = len(str(headers[col_idx - 1]))
        for r in range(2, ws.max_row + 1):
            mx = max(mx, len(str(ws.cell(r, col_idx).value or "")))
        ws.column_dimensions[letter].width = min(mx + 2, 48)


def _add_readme_sheet(wb: Workbook, lines: list[str]) -> None:
    ws = wb.create_sheet("README", 0)
    for i, text in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 100


def generate_workbook_for_dataset(
    dataset_slug: str,
    *,
    with_sample: bool = True,
    include_notes: bool = True,
) -> Workbook:
    definition = get_dataset(dataset_slug)
    wb = Workbook()
    wb.remove(wb.active)
    if include_notes:
        if definition.template_mode == "universal_long":
            _add_readme_sheet(
                wb,
                [
                    "Template universal — satu format untuk berbagai sumber (PLN, BI, BPJS, dll.).",
                    "Lembar data: kolom wajib baris 1 — nama_dataset | indikator | periode | nilai.",
                    "nama_dataset: nama kelompok data Anda (bebas teks).",
                    "indikator: satu kolom; beberapa dimensi dipisah dengan | (contoh: Kelompok | Metrik | Detail).",
                    "periode: gunakan salah satu format — YYYY, YYYY-MM, YYYY-Q1 / Q1-YYYY, atau tanggal YYYY-MM-DD.",
                    "nilai: angka. Kosongkan baris contoh jika tidak dipakai.",
                    "Form unggah web: pilih Flow/Stock dan Bulanan/Triwulanan/Tahunan sesuai konteks; metadata itu berbeda dari format tanggal di kolom periode.",
                ],
            )
        else:
            _add_readme_sheet(
                wb,
                [
                    f"Template long-format: {definition.label}",
                    f"Sheet sumber BI: {definition.source_sheet!r}",
                    f"Mode waktu: {definition.time_period_mode}; tipe tabel: {definition.table_type}.",
                    "Baris 1 = header wajib. Hapus baris contoh lalu isi data. Satuan nilai mengikuti publikasi BI.",
                ],
            )
    _write_long_sheet(wb, definition, with_sample=with_sample)
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_universal_template_file_response() -> Response:
    """Same workbook as ``generate_workbook_for_dataset('universal')`` with stable filename."""
    from flask import Response

    wb = generate_workbook_for_dataset("universal", with_sample=True, include_notes=True)
    data = workbook_to_bytes(wb)
    fname = _attachment_filename("template_universal.xlsx")
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def build_template_file_response(dataset_slug: str) -> Response:
    from flask import Response

    if (dataset_slug or "").strip().lower() == "universal":
        return build_universal_template_file_response()

    definition = get_dataset(dataset_slug)
    wb = generate_workbook_for_dataset(dataset_slug, with_sample=True, include_notes=True)
    data = workbook_to_bytes(wb)
    fname = _attachment_filename(f"template_rekap_{definition.slug}.xlsx")
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _attachment_filename(name: str) -> str:
    base = re.sub(r'[^A-Za-z0-9._-]+', "_", name).strip("_") or "template.xlsx"
    return base[:120]


def build_multi_dataset_reference_workbook(*, with_sample: bool = True) -> Workbook:
    """Workbook multi-tab + README (untuk `static/templates/rekap_dataset_long_templates.xlsx`)."""
    wb = Workbook()
    wb.remove(wb.active)
    _add_readme_sheet(
        wb,
        [
            "Long-format upload templates untuk dataset REKAP BI.",
            "Satu tab = satu dataset (slug). Baris 1 = header wajib.",
            "Bulan = 1–12; triwulan (ecommerce) = 1–4 (= Tw I–Tw IV).",
            "jenis_nilai (ATM / kartu kredit / uang elektronik): volume | nominal (huruf kecil).",
            "Sheet sumber kartu kredit di Excel BI: 'Kartu kredit ' (spasi akhir).",
        ],
    )
    for d in iter_datasets():
        if d.slug == "universal":
            continue
        _write_long_sheet(wb, d, with_sample=with_sample)
    return wb


def write_multi_dataset_reference_workbook(path: str | bytes, *, with_sample: bool = True) -> None:
    wb = build_multi_dataset_reference_workbook(with_sample=with_sample)
    wb.save(path)
