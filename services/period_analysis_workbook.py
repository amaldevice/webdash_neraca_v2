# -*- coding: utf-8 -*-
"""OpenPyXL workbook construction and cell styling for period analysis export."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _calc_avg_change(comparison_list: list[dict]) -> str:
    if not comparison_list:
        return "N/A"
    try:
        avg = sum(item.get("change_percent", 0) for item in comparison_list) / len(comparison_list)
        return f"{avg:+.2f}%"
    except (TypeError, ZeroDivisionError):
        return "N/A"


@dataclass(frozen=True)
class _PeriodAnalysisStyles:
    header_font: Font
    header_fill: PatternFill
    center_align: Alignment
    thin_border: Border
    green_fill: PatternFill
    red_fill: PatternFill
    green_font: Font
    red_font: Font
    title_font: Font
    subtitle_font: Font
    section_font: Font
    ytd_accum_fill: PatternFill


def _period_analysis_styles() -> _PeriodAnalysisStyles:
    return _PeriodAnalysisStyles(
        header_font=Font(bold=True, color="FFFFFF", size=11),
        header_fill=PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        center_align=Alignment(horizontal="center", vertical="center"),
        thin_border=Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
        green_fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        red_fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        green_font=Font(color="006100", bold=True),
        red_font=Font(color="9C0006", bold=True),
        title_font=Font(bold=True, size=16, color="366092"),
        subtitle_font=Font(italic=True, size=10, color="666666"),
        section_font=Font(bold=True, size=12, color="366092"),
        ytd_accum_fill=PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
    )


def _apply_header_style(cell, styles: _PeriodAnalysisStyles) -> None:
    cell.font = styles.header_font
    cell.fill = styles.header_fill
    cell.alignment = styles.center_align
    cell.border = styles.thin_border


def _apply_data_style(
    cell,
    styles: _PeriodAnalysisStyles,
    *,
    center: bool = False,
    fill=None,
    font=None,
) -> None:
    cell.border = styles.thin_border
    if center:
        cell.alignment = styles.center_align
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font


def _build_dashboard_sheet(
    wb: Workbook,
    indicator: str,
    results: dict[str, Any],
    styles: _PeriodAnalysisStyles,
) -> None:
    has_mm = bool(results.get("monthly_comparison") and len(results["monthly_comparison"]) > 0)
    has_qq = bool(results.get("quarterly_comparison") and len(results["quarterly_comparison"]) > 0)
    has_yy = bool(results.get("yearly_comparison") and len(results["yearly_comparison"]) > 0)
    has_ytd = bool(results.get("ytd_comparison") and len(results["ytd_comparison"]) > 0)
    has_cc = bool(results.get("current_to_current") and len(results["current_to_current"]) > 0)

    row_data = [
        ("M ke M (Bulanan)", len(results.get("monthly_comparison", [])), _calc_avg_change(results.get("monthly_comparison", [])), "✓" if has_mm else "✗"),
        ("Q ke Q (Kuartal)", len(results.get("quarterly_comparison", [])), _calc_avg_change(results.get("quarterly_comparison", [])), "✓" if has_qq else "✗"),
        ("Y ke Y (Tahunan)", len(results.get("yearly_comparison", [])), _calc_avg_change(results.get("yearly_comparison", [])), "✓" if has_yy else "✗"),
        ("YTD (Akumulasi)", sum(len(y.get("monthly_ytd", [])) for y in results.get("ytd_comparison", [])), "-", "✓" if has_ytd else "✗"),
        ("C ke C (YoY)", len(results.get("current_to_current", [])), _calc_avg_change(results.get("current_to_current", [])), "✓" if has_cc else "✗"),
    ]

    ws = wb.active
    ws.title = "Dashboard"

    ws["A1"] = f"📊 Analisis Periode: {indicator}"
    ws["A1"].font = styles.title_font
    ws.merge_cells("A1:D1")

    ws["A2"] = (
        f"Tahun Analisis: {results.get('analysis_year', 'Semua Tahun')} | "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws["A2"].font = styles.subtitle_font
    ws.merge_cells("A2:D2")

    ws["A4"] = "📈 Ringkasan Data"
    ws["A4"].font = styles.section_font
    ws.merge_cells("A4:D4")

    for col, header in enumerate(["Metrik", "Jumlah Data", "Rata-rata Perubahan", "Status"], 1):
        _apply_header_style(ws.cell(row=5, column=col, value=header), styles)

    for idx, (metric, count, avg_change, status) in enumerate(row_data, start=6):
        _apply_data_style(ws.cell(row=idx, column=1, value=metric), styles)
        _apply_data_style(ws.cell(row=idx, column=2, value=count), styles, center=True)
        _apply_data_style(ws.cell(row=idx, column=3, value=avg_change), styles, center=True)
        _apply_data_style(ws.cell(row=idx, column=4, value=status), styles, center=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10


def _build_comparison_sheet(
    wb: Workbook,
    sheet_name: str,
    headers: list[str],
    data_items: list[dict[str, Any]],
    styles: _PeriodAnalysisStyles,
    indicator: str,
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"{sheet_name} - {indicator}"
    ws["A1"].font = styles.section_font
    ws.merge_cells("A1:E1")

    for col, header in enumerate(headers, 1):
        _apply_header_style(ws.cell(row=3, column=col, value=header), styles)

    for idx, item in enumerate(data_items, start=4):
        _apply_data_style(ws.cell(row=idx, column=1, value=item["period"]), styles, center=True)
        _apply_data_style(ws.cell(row=idx, column=2, value=item["current_value"]), styles, center=True)
        _apply_data_style(ws.cell(row=idx, column=3, value=item["previous_value"]), styles, center=True)

        change_cell = ws.cell(row=idx, column=4, value=f"{item['change_percent']}%")
        status_cell = ws.cell(row=idx, column=5, value=item["change_type"].capitalize())

        _apply_data_style(change_cell, styles, center=True)
        _apply_data_style(status_cell, styles, center=True)

        change_type = item["change_type"]
        if change_type == "increase":
            _apply_data_style(change_cell, styles, center=True, fill=styles.green_fill, font=styles.green_font)
            _apply_data_style(status_cell, styles, center=True, fill=styles.green_fill, font=styles.green_font)
        elif change_type == "decrease":
            _apply_data_style(change_cell, styles, center=True, fill=styles.red_fill, font=styles.red_font)
            _apply_data_style(status_cell, styles, center=True, fill=styles.red_fill, font=styles.red_font)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12


def _build_ytd_sheet(
    wb: Workbook,
    results: dict[str, Any],
    styles: _PeriodAnalysisStyles,
    indicator: str,
) -> None:
    ws = wb.create_sheet("YTD")
    ws["A1"] = f"YTD Analysis - {indicator}"
    ws["A1"].font = styles.section_font
    ws.merge_cells("A1:D1")

    current_row = 3
    for year_data in results.get("ytd_comparison", []):
        year = year_data.get("year", "Unknown")
        ws.cell(row=current_row, column=1, value=f"Tahun {year}").font = Font(
            bold=True,
            size=11,
            color="366092",
        )
        current_row += 1

        ytd_headers = ["Periode", "Nilai Bulanan", "YTD Akumulasi", "Status"]
        for col, header in enumerate(ytd_headers, 1):
            _apply_header_style(ws.cell(row=current_row, column=col, value=header), styles)
        current_row += 1

        monthly_data = year_data.get("monthly_ytd", [])
        for item in monthly_data:
            _apply_data_style(ws.cell(row=current_row, column=1, value=item["period"]), styles, center=True)
            _apply_data_style(ws.cell(row=current_row, column=2, value=item.get("monthly_value", "-")), styles, center=True)
            ytd_cell = ws.cell(row=current_row, column=3, value=item["ytd_value"])
            _apply_data_style(ytd_cell, styles, center=True, fill=styles.ytd_accum_fill)
            _apply_data_style(ws.cell(row=current_row, column=4, value="Akumulasi"), styles, center=True)
            current_row += 1

        current_row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14


def _build_metadata_sheet(wb: Workbook, indicator: str, results: dict[str, Any], styles: _PeriodAnalysisStyles) -> None:
    ws = wb.create_sheet("Metadata")

    ws["A1"] = "📋 Informasi Analisis"
    ws["A1"].font = styles.title_font
    ws.merge_cells("A1:C1")

    metadata_items = [
        ("Indikator", indicator),
        ("Tahun Analisis", results.get("analysis_year", "Semua Tahun")),
        ("Tanggal Export", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Jumlah Periode M-M", str(len(results.get("monthly_comparison", [])))),
        ("Jumlah Periode Q-Q", str(len(results.get("quarterly_comparison", [])))),
        ("Jumlah Periode Y-Y", str(len(results.get("yearly_comparison", [])))),
        ("Jumlah Data C-C", str(len(results.get("current_to_current", [])))),
        ("", ""),
        ("Catatan", "Data ini dihasilkan oleh Sistem Data BPS"),
        ("Disclaimer", "Data untuk analisis internal. Validasi ke sumber resmi BPS sebelum dipublikasikan."),
    ]

    for idx, (label, value) in enumerate(metadata_items, start=3):
        label_cell = ws.cell(row=idx, column=1, value=label)
        label_cell.font = Font(bold=bool(label))
        value_cell = ws.cell(row=idx, column=2, value=value)
        if label in ("Catatan", "Disclaimer"):
            value_cell.font = Font(italic=True, size=9)
            ws.merge_cells(f"B{idx}:C{idx}")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20


def _build_current_to_current_sheet(
    wb: Workbook,
    results: dict[str, Any],
    styles: _PeriodAnalysisStyles,
    indicator: str,
) -> None:
    ws = wb.create_sheet("C ke C")
    ws["A1"] = f"C ke C (Current vs Previous Year) - {indicator}"
    ws["A1"].font = styles.section_font
    ws.merge_cells("A1:E1")

    cc_headers = ["Periode", "Nilai Tahun Ini", "Nilai Tahun Lalu", "Perubahan (%)", "Status"]
    for col, header in enumerate(cc_headers, 1):
        _apply_header_style(ws.cell(row=3, column=col, value=header), styles)

    for idx, item in enumerate(results["current_to_current"], start=4):
        _apply_data_style(ws.cell(row=idx, column=1, value=item["period"]), styles, center=True)
        _apply_data_style(ws.cell(row=idx, column=2, value=item["current_value"]), styles, center=True)
        _apply_data_style(ws.cell(row=idx, column=3, value=item["previous_year_value"]), styles, center=True)

        change_cell = ws.cell(row=idx, column=4, value=f"{item['change_percent']}%")
        status_cell = ws.cell(row=idx, column=5, value=item["change_type"].capitalize())
        _apply_data_style(change_cell, styles, center=True)
        _apply_data_style(status_cell, styles, center=True)

        change_type = item["change_type"]
        if change_type == "increase":
            _apply_data_style(change_cell, styles, center=True, fill=styles.green_fill, font=styles.green_font)
            _apply_data_style(status_cell, styles, center=True, fill=styles.green_fill, font=styles.green_font)
        elif change_type == "decrease":
            _apply_data_style(change_cell, styles, center=True, fill=styles.red_fill, font=styles.red_font)
            _apply_data_style(status_cell, styles, center=True, fill=styles.red_fill, font=styles.red_font)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12


def build_period_analysis_workbook(indicator: str, results: dict[str, Any]) -> Workbook:
    """
    Populate a multi-sheet workbook from `calculate_period_comparisons` result dict.
    Caller serializes to bytes / HTTP response.
    """
    wb = Workbook()
    styles = _period_analysis_styles()

    has_mm = bool(results.get("monthly_comparison") and len(results["monthly_comparison"]) > 0)
    has_qq = bool(results.get("quarterly_comparison") and len(results["quarterly_comparison"]) > 0)
    has_yy = bool(results.get("yearly_comparison") and len(results["yearly_comparison"]) > 0)
    has_ytd = bool(results.get("ytd_comparison") and len(results["ytd_comparison"]) > 0)
    has_cc = bool(results.get("current_to_current") and len(results["current_to_current"]) > 0)

    _build_dashboard_sheet(wb, indicator, results, styles)

    if has_mm:
        _build_comparison_sheet(
            wb,
            "M to M",
            ["Bulan", "Nilai", "Sebelumnya", "Perubahan", "Status"],
            results["monthly_comparison"],
            styles,
            indicator,
        )
    if has_qq:
        _build_comparison_sheet(
            wb,
            "Q ke Q",
            ["Kuartal", "Nilai", "Sebelumnya", "Perubahan", "Status"],
            results["quarterly_comparison"],
            styles,
            indicator,
        )
    if has_yy:
        _build_comparison_sheet(
            wb,
            "Y ke Y",
            ["Tahun", "Nilai", "Sebelumnya", "Perubahan", "Status"],
            results["yearly_comparison"],
            styles,
            indicator,
        )
    if has_ytd:
        _build_ytd_sheet(wb, results, styles, indicator)
    if has_cc:
        _build_current_to_current_sheet(wb, results, styles, indicator)

    _build_metadata_sheet(wb, indicator, results, styles)
    return wb
