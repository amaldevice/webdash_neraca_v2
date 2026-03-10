# -*- coding: utf-8 -*-
import csv
import io
import json
import os
import sqlite3
from datetime import datetime

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.utils import PlotlyJSONEncoder

from flask import Flask, Response, render_template, request, redirect, url_for, flash, jsonify
from werkzeug.utils import secure_filename

from aggregator import fetch_aggregated_summary, refresh_aggregated_summary
from excel_parser import parse_excel
from models import (
    init_db, insert_entries, query_data_entries, get_total_entries_count,
    delete_data_entry, delete_data_by_filter, update_data_entry, insert_single_entry,
    get_filter_options, update_data_entry_full, get_unique_indicators,
    bulk_delete_entries, bulk_update_entries, calculate_period_comparisons
)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
ALLOWED_EXTENSIONS = {"xls", "xlsx"}
ALLOWED_DATA_TYPES = {"flow", "stock"}
ALLOWED_TIME_PERIODS = {"monthly", "quarterly", "yearly"}

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = os.path.join(BASE_DIR, "uploads")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024
app.config["SECRET_KEY"] = os.environ.get("FLASK_SECRET_KEY", "change-me-for-production")

init_db()

def allowed_file(filename: str) -> bool:
    """Validate file extensions for Excel uploads."""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def validate_metadata(data_type: str, time_period: str) -> list[str]:
    errors = []
    if data_type.lower() not in ALLOWED_DATA_TYPES:
        errors.append("Jenis data tidak valid.")
    if time_period.lower() not in ALLOWED_TIME_PERIODS:
        errors.append("Periode tidak valid.")
    return errors


@app.route("/", methods=["GET"])
def landing_page():
    summary = fetch_aggregated_summary()
    return render_template("landing.html", summary=summary)


@app.route("/upload", methods=["GET", "POST"])
def upload_data():
    if request.method == "POST":
        uploader = request.form.get("uploader", "").strip()
        version = request.form.get("version", "").strip()
        data_type = request.form.get("data_type", "flow").strip()
        time_period = request.form.get("time_period", "monthly").strip()
        file = request.files.get("excel_file")
        errors = []

        if not uploader:
            errors.append("Nama pengunggah wajib diisi.")
        if not version:
            errors.append("Versi wajib diisi.")
        if not file or not allowed_file(file.filename):
            errors.append("Harus mengunggah file Excel (.xls/.xlsx).")
        errors.extend(validate_metadata(data_type, time_period))

        if errors:
            for error in errors:
                flash(error, "error")
        else:
            filename = secure_filename(file.filename)
            os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
            destination = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            file.save(destination)
            entries = parse_excel(destination, uploader, version, data_type, time_period)
            if not entries:
                flash("File Excel tidak berisi data yang valid.", "error")
            else:
                try:
                    insert_entries(entries)
                    refresh_aggregated_summary()
                    flash(f"{len(entries)} baris data berhasil disimpan.", "success")
                except sqlite3.IntegrityError as e:
                    # Handle database constraint violations
                    error_msg = str(e)
                    if "UNIQUE constraint failed" in error_msg:
                        flash("Data dengan pengunggah, versi, dan indikator yang sama sudah ada. Gunakan versi yang berbeda atau periksa file Excel Anda.", "error")
                    else:
                        flash(f"Terjadi kesalahan database: {error_msg}", "error")
                except Exception as e:
                    # Handle other unexpected errors
                    flash(f"Terjadi kesalahan saat menyimpan data: {str(e)}", "error")
            return redirect(url_for("upload_data"))

    return render_template("upload.html", mode="upload")


@app.route("/manual", methods=["GET", "POST"])
def manual_input():
    if request.method == "POST":
        uploader = request.form.get("uploader", "").strip()
        version = request.form.get("version", "").strip()
        data_type = request.form.get("data_type", "flow").strip()
        time_period = request.form.get("time_period", "monthly").strip()
        period_date = request.form.get("period_date", "").strip()
        indicator = request.form.get("indicator", "").strip()
        value = request.form.get("value", "").strip()

        if not uploader or not version or not indicator or not value or not period_date:
            flash("Semua kolom metadata dan data wajib diisi.", "error")
        else:
            validation_errors = validate_metadata(data_type, time_period)
            if validation_errors:
                for err in validation_errors:
                    flash(err, "error")
            else:
                manual_entry = _build_manual_entry(
                    uploader, version, data_type, time_period, period_date, indicator, value
                )
                if manual_entry is None:
                    flash("Nilai indikator tidak valid.", "error")
                else:
                    insert_entries([manual_entry])
                    refresh_aggregated_summary()
                    flash("Entri manual berhasil dicatat dan disimpan.", "success")
                    return redirect(url_for("manual_input"))

    return render_template("upload.html", mode="manual")


def _parse_period_date(time_period: str, period_date: str) -> tuple[int | None, int | None, int | None]:
    """Parse period_date string into year, month, quarter based on time_period format."""
    try:
        if time_period.lower() == 'monthly':
            # Format: YYYY-MM
            if '-' in period_date and len(period_date.split('-')) == 2:
                year_str, month_str = period_date.split('-')
                year = int(year_str)
                month = int(month_str)
                quarter = (month - 1) // 3 + 1  # Calculate quarter from month
                return year, month, quarter
        elif time_period.lower() == 'quarterly':
            # Format: YYYY-Q1/Q2/Q3/Q4
            if '-Q' in period_date:
                year_str, quarter_str = period_date.split('-Q')
                year = int(year_str)
                quarter = int(quarter_str)
                return year, None, quarter
        elif time_period.lower() == 'yearly':
            # Format: YYYY
            year = int(period_date)
            return year, None, None
    except (ValueError, IndexError):
        pass

    return None, None, None


def _build_manual_entry(
    uploader: str,
    version: str,
    data_type: str,
    time_period: str,
    period_date: str,
    indicator: str,
    value: str,
) -> dict | None:
    try:
        parsed_value = float(value)
    except ValueError:
        return None

    # Parse period_date based on time_period
    year, month, quarter = _parse_period_date(time_period, period_date)

    return {
        "uploader_name": uploader,
        "version": version,
        "template_type": "manual",
        "data_type": data_type.lower() or "flow",
        "time_period": time_period.lower() or "monthly",
        "indicator_name": indicator,
        "value": parsed_value,
        "unit": None,
        "region_code": None,
        "year": year,
        "month": month,
        "quarter": quarter,
        "created_at": datetime.utcnow().isoformat(),
    }


def get_available_years() -> list[int]:
    """Fetch distinct year values for aggregated UI filters."""
    db_path = os.path.join(BASE_DIR, "data.db")
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT year
            FROM data_entries
            WHERE year IS NOT NULL
            ORDER BY year DESC
            """
        ).fetchall()
    years: list[int] = []
    for row in rows:
        raw_year = row[0]
        if raw_year is None:
            continue
        try:
            years.append(int(raw_year))
        except (TypeError, ValueError):
            continue
    return years


def _get_range_params(source, start_key: str = "start_period", end_key: str = "end_period") -> tuple[str | None, str | None]:
    """Extract optional period-range parameters from request-like inputs."""
    start_period = source.get(start_key, "", type=str)
    end_period = source.get(end_key, "", type=str)
    return (
        start_period.strip() if start_period is not None else None,
        end_period.strip() if end_period is not None else None,
    )


@app.route("/preview-data", methods=["GET"])
def preview_data():
    # Get filter parameters
    data_type = request.args.get("data_type", "")
    time_period = request.args.get("time_period", "")
    uploader = request.args.get("uploader", "")
    indicator = request.args.get("indicator", "")
    period_start, period_end = _get_range_params(request.args)

    # Get pagination parameters
    page = int(request.args.get("page", 1))
    limit = int(request.args.get("limit", 20))  # Default 20 rows

    # Validate limit options
    if limit not in [5, 10, 15, 20, 30, 50, 100]:
        limit = 20

    # Calculate offset for pagination
    offset = (page - 1) * limit

    summary = fetch_aggregated_summary()

    # Query data with all filters at SQL level
    entries = query_data_entries(
        data_type=data_type or None,
        time_period=time_period or None,
        uploader=uploader or None,
        indicator=indicator or None,
        period_start=period_start,
        period_end=period_end,
        limit=limit,
        offset=offset
    )

    # Get total count for pagination (considering all filters)
    total_entries = get_total_entries_count(
        data_type=data_type or None,
        time_period=time_period or None,
        uploader=uploader or None,
        indicator=indicator or None,
        period_start=period_start,
        period_end=period_end
    )

    total_pages = (total_entries + limit - 1) // limit  # Ceiling division

    filters = {
        "data_type": data_type,
        "time_period": time_period,
        "uploader": uploader,
        "indicator": indicator,
        "start_period": period_start,
        "end_period": period_end,
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "total_entries": total_entries
    }

    # Get filter options for autocomplete
    filter_options = get_filter_options()

    return render_template(
        "preview.html", summary=summary, entries=entries, filters=filters, filter_options=filter_options
    )


@app.route("/export", methods=["GET"])
def export_data():
    fmt = request.args.get("format", "csv").lower()
    data_type = request.args.get("data_type")
    time_period = request.args.get("time_period")
    uploader = request.args.get("uploader")
    indicator = request.args.get("indicator")
    period_start, period_end = _get_range_params(request.args)
    entries = query_data_entries(
        data_type=data_type,
        time_period=time_period,
        uploader=uploader,
        indicator=indicator,
        period_start=period_start,
        period_end=period_end,
        limit=1000
    )
    if fmt == "excel":
        # Ensure proper column structure even for empty data
        columns = ["id", "uploader_name", "version", "indicator_name", "value", "data_type", "time_period", "created_at", "year", "month", "quarter"]
        df = pd.DataFrame(entries, columns=columns) if entries else pd.DataFrame(columns=columns)
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        headers = {
            "Content-Disposition": "attachment; filename=raw-data.xlsx",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        return Response(output.getvalue(), headers=headers)

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    headers = ["id", "uploader_name", "version", "indicator_name", "value", "data_type", "time_period", "created_at"]
    writer.writerow(headers)
    for row in entries:
        writer.writerow(
            [
                row["id"],
                row["uploader_name"],
                row["version"],
                row["indicator_name"],
                row["value"],
                row["data_type"],
                row["time_period"],
                row["created_at"],
            ]
        )

    response = Response(buffer.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=raw-data.csv"
    return response


@app.route("/aggregated", methods=["GET"])
def aggregated_summary():
    summary = fetch_aggregated_summary()
    # Get unique indicators for filter dropdown
    indicators = get_unique_indicators()
    summary_dict = summary if isinstance(summary, dict) else {}
    available_years = summary_dict.get("available_years")
    if not isinstance(available_years, list) or not available_years:
        available_years = get_available_years()
    summary_with_indicators = dict(summary_dict)
    summary_with_indicators['indicators'] = indicators
    summary_with_indicators['available_years'] = available_years
    return render_template("aggregated.html", summary=summary_with_indicators)


@app.route("/generate-plot", methods=["POST"])
def generate_plot():
    indicator = request.form.get("indicator_filter", "").strip()
    time_range = request.form.get("time_range", "all")
    period_start, period_end = _get_range_params(request.form, start_key="period_start", end_key="period_end")

    if not indicator:
        return jsonify({"error": "Pilih indikator terlebih dahulu"})

    if time_range == "all":
        range_start = None
        range_end = None
    else:
        normalized_year = time_range.strip()
        range_start = normalized_year
        range_end = normalized_year

    start_period = period_start or range_start
    end_period = period_end or range_end

    fig_json = generate_indicator_line_chart(indicator, time_range, start_period, end_period)
    return jsonify({"plot_json": fig_json})


@app.route("/generate-period-analysis", methods=["POST"])
def generate_period_analysis():
    indicator = request.form.get("indicator", "").strip()
    analysis_year = request.form.get("year", "").strip()
    period_start, period_end = _get_range_params(request.form, start_key="period_start", end_key="end_period")

    if not indicator:
        return jsonify({"error": "Pilih indikator terlebih dahulu"})

    results = calculate_period_comparisons(indicator, analysis_year or None, period_start, period_end)

    if "error" in results:
        return jsonify({"error": results["error"]})

    return jsonify({"analysis": results})


def generate_indicator_line_chart(
    indicator,
    time_range="all",
    period_start: str | None = None,
    period_end: str | None = None
):
    """Generate line chart for a specific indicator with time filtering"""

    # Get data for specific indicator (with optional period filter)
    entries = query_data_entries(
        indicator=indicator,
        limit=1000,
        period_start=period_start,
        period_end=period_end,
    )

    if not entries:
        return "<div class='no-data'>Tidak ada data untuk membuat plot</div>"

    # Convert to DataFrame
    df = pd.DataFrame(entries)

    # Filter by selected indicator
    df_filtered = df[df['indicator_name'] == indicator].copy()

    if df_filtered.empty:
        return f"<div class='error'>Tidak ada data untuk indikator '{indicator}'</div>"

    if time_range != "all":
        # Keep backward-compatible explicit time_range filtering if provided
        if time_range and 'tanggal_data' in df_filtered.columns:
            df_filtered = df_filtered[df_filtered['tanggal_data'].str.startswith(time_range)]

    # Sort by date for proper line chart
    if 'tanggal_data' in df_filtered.columns:
        df_filtered = df_filtered.sort_values('tanggal_data')
    else:
        df_filtered = df_filtered.sort_index()

    # Create line chart
    fig = px.line(
        df_filtered,
        x='tanggal_data' if 'tanggal_data' in df_filtered.columns else df_filtered.index,
        y='value',
        title=f'Tren {indicator} ({time_range.replace("all", "Semua Periode")})',
        labels={
            'tanggal_data': 'Periode',
            'value': 'Nilai',
            'x': 'Periode'
        },
        markers=True  # Add markers for better visibility
    )

    # Update line appearance
    fig.update_traces(
        line=dict(width=3),
        marker=dict(size=8)
    )

    # Update layout
    fig.update_layout(
        template='plotly_white',
        font=dict(size=12),
        margin=dict(l=20, r=20, t=50, b=20),
        xaxis_title="Periode Waktu",
        yaxis_title="Nilai",
        showlegend=False  # Hide legend since only one line
    )

    # Convert to HTML
    # NOTE: Using include_plotlyjs=True inlines the entire Plotly.js bundle into
    # every response (very large). This can make the /generate-plot response
    # feel like it "doesn't work" on slower machines/browsers.
    # Using CDN keeps responses small and usually fixes the UX.
    fig.update_layout(height=480)

    # Return JSON instead of HTML-with-<script>. Scripts injected via innerHTML are
    # not reliably executed by browsers, which can result in a blank plot area.
    # Client will render using Plotly.newPlot().
    return fig.to_json()


@app.route("/data-management", methods=["GET", "POST"])
def data_management():
    # Handle filters
    filter_source = request.values
    data_type = filter_source.get("data_type", "")
    time_period = filter_source.get("time_period", "")
    uploader = filter_source.get("uploader", "")
    indicator = filter_source.get("indicator", "")
    period_start, period_end = _get_range_params(filter_source)

    # Handle pagination parameters
    page = int(request.args.get("page", 1))
    limit = int(request.args.get("limit", 20))  # Default 20 rows for management

    # Validate limit options for data management
    if limit not in [5, 10, 15, 20, 30, 50, 100]:
        limit = 20

    # Calculate offset for pagination
    offset = (page - 1) * limit

    # Handle CRUD operations
    if request.method == "POST":
        action = request.form.get("action")

        if action == "delete_single":
            entry_id = request.form.get("entry_id")
            if entry_id:
                delete_data_entry(entry_id)
                flash("Data berhasil dihapus.", "success")

        elif action == "delete_by_filter":
            deleted_count = delete_data_by_filter(
                data_type=data_type or None,
                time_period=time_period or None,
                uploader=uploader or None,
                indicator=indicator or None,
                period_start=period_start,
                period_end=period_end
            )
            flash(f"{deleted_count} data berhasil dihapus berdasarkan filter.", "success")

        elif action == "update":
            entry_id = request.form.get("entry_id")
            update_uploader = request.form.get("update_uploader", "").strip()
            update_version = request.form.get("update_version", "").strip()
            update_indicator = request.form.get("update_indicator", "").strip()
            update_value = request.form.get("update_value", "").strip()
            update_data_type = request.form.get("update_data_type", "").strip()
            update_time_period = request.form.get("update_time_period", "").strip()

            if entry_id and all([update_uploader, update_version, update_indicator, update_value]):
                try:
                    update_data_entry_full(entry_id, {
                        "uploader_name": update_uploader,
                        "version": update_version,
                        "indicator_name": update_indicator,
                        "value": float(update_value),
                        "data_type": update_data_type,
                        "time_period": update_time_period
                    })
                    flash("Data berhasil diperbarui.", "success")
                    # Clear cache after update
                    refresh_aggregated_summary()
                except ValueError:
                    flash("Nilai harus berupa angka.", "error")
                except Exception as e:
                    flash(f"Terjadi kesalahan saat memperbarui data: {str(e)}", "error")

        elif action == "bulk_delete":
            selected_ids = request.form.getlist("selected_ids[]")
            if selected_ids:
                deleted_count = bulk_delete_entries(selected_ids)
                flash(f"{deleted_count} data berhasil dihapus.", "success")
                refresh_aggregated_summary()
            else:
                flash("Tidak ada data yang dipilih untuk dihapus.", "error")

        elif action == "bulk_update":
            selected_ids = request.form.getlist("selected_ids[]")
            if not selected_ids:
                flash("Tidak ada data yang dipilih untuk diperbarui.", "error")
            else:
                # Get update values (only non-empty fields will be updated)
                updates = {}
                update_uploader = request.form.get("bulk_update_uploader", "").strip()
                update_version = request.form.get("bulk_update_version", "").strip()
                update_data_type = request.form.get("bulk_update_data_type", "").strip()
                update_time_period = request.form.get("bulk_update_time_period", "").strip()
                update_value = request.form.get("bulk_update_value", "").strip()

                if update_uploader:
                    updates["uploader_name"] = update_uploader
                if update_version:
                    updates["version"] = update_version
                if update_data_type:
                    updates["data_type"] = update_data_type
                if update_time_period:
                    updates["time_period"] = update_time_period
                if update_value:
                    try:
                        updates["value"] = float(update_value)
                    except ValueError:
                        flash("Nilai harus berupa angka.", "error")
                        return redirect(url_for("data_management", data_type=data_type, time_period=time_period,
                                              uploader=uploader, indicator=indicator,
                                              start_period=period_start, end_period=period_end))

                if updates:
                    updated_count = bulk_update_entries(selected_ids, updates)
                    flash(f"{updated_count} data berhasil diperbarui.", "success")
                    refresh_aggregated_summary()
                else:
                    flash("Tidak ada kolom yang diisi untuk diperbarui.", "error")

        elif action == "insert":
            insert_uploader = request.form.get("insert_uploader", "").strip()
            insert_version = request.form.get("insert_version", "").strip()
            insert_data_type = request.form.get("insert_data_type", "").strip()
            insert_time_period = request.form.get("insert_time_period", "").strip()
            insert_period_date = request.form.get("insert_period_date", "").strip()
            insert_indicator = request.form.get("insert_indicator", "").strip()
            insert_value = request.form.get("insert_value", "").strip()

            if all([insert_uploader, insert_version, insert_indicator, insert_value, insert_period_date]):
                try:
                    insert_single_entry(
                        uploader=insert_uploader,
                        version=insert_version,
                        data_type=insert_data_type,
                        time_period=insert_time_period,
                        period_date=insert_period_date,
                        indicator=insert_indicator,
                        value=float(insert_value)
                    )
                    flash("Data baru berhasil ditambahkan.", "success")
                    # Clear cache after insert
                    refresh_aggregated_summary()
                except ValueError:
                    flash("Nilai harus berupa angka.", "error")
            else:
                flash("Semua kolom wajib diisi.", "error")

        return redirect(url_for(
            "data_management",
            data_type=data_type,
            time_period=time_period,
            uploader=uploader,
            indicator=indicator,
            start_period=period_start,
            end_period=period_end
        ))

    # Get data for display with proper pagination
    entries = query_data_entries(
        data_type=data_type or None,
        time_period=time_period or None,
        uploader=uploader or None,
        indicator=indicator or None,
        period_start=period_start,
        period_end=period_end,
        limit=limit,
        offset=offset
    )

    # Get total count for pagination
    total_entries = get_total_entries_count(
        data_type=data_type or None,
        time_period=time_period or None,
        uploader=uploader or None,
        indicator=indicator or None,
        period_start=period_start,
        period_end=period_end
    )

    total_pages = (total_entries + limit - 1) // limit  # Ceiling division

    filters = {
        "data_type": data_type,
        "time_period": time_period,
        "uploader": uploader,
        "indicator": indicator,
        "start_period": period_start,
        "end_period": period_end,
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "total_entries": total_entries
    }

    # Get filter options for autocomplete
    filter_options = get_filter_options()

    return render_template("data_management.html", entries=entries, filters=filters, filter_options=filter_options)


@app.route("/export-period-analysis", methods=["POST"])
def export_period_analysis_excel():
    """Export period analysis to Excel (Phase 2: Complete with all sheets + conditional formatting)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference

    indicator = request.form.get("indicator", "").strip()
    analysis_year = request.form.get("year", "").strip()
    period_start, period_end = _get_range_params(request.form, start_key="period_start", end_key="end_period")

    if not indicator:
        flash("Pilih indikator terlebih dahulu.", "error")
        return redirect(url_for("aggregated_summary"))

    # Convert year to int if it's numeric, otherwise None for all years
    year_param = None
    if analysis_year and analysis_year.isdigit():
        year_param = int(analysis_year)

    # Get analysis data
    results = calculate_period_comparisons(indicator, year_param, period_start, period_end)

    if "error" in results:
        flash(results["error"], "error")
        return redirect(url_for("aggregated_summary"))

    # Create workbook
    wb = Workbook()

    # Define styles - Enhanced for Phase 2
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Conditional formatting colors (background)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red
    green_font = Font(color="006100", bold=True)  # Dark green
    red_font = Font(color="9C0006", bold=True)    # Dark red
    
    title_font = Font(bold=True, size=16, color="366092")
    subtitle_font = Font(italic=True, size=10, color="666666")
    section_font = Font(bold=True, size=12, color="366092")

    # Check what data exists
    has_mm = bool(results.get('monthly_comparison') and len(results['monthly_comparison']) > 0)
    has_qq = bool(results.get('quarterly_comparison') and len(results['quarterly_comparison']) > 0)
    has_yy = bool(results.get('yearly_comparison') and len(results['yearly_comparison']) > 0)
    has_ytd = bool(results.get('ytd_comparison') and len(results['ytd_comparison']) > 0)
    has_cc = bool(results.get('current_to_current') and len(results['current_to_current']) > 0)

    # ========== SHEET 1: DASHBOARD (Enhanced) ==========
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"

    # Title
    ws_dashboard['A1'] = f"📊 Analisis Periode: {indicator}"
    ws_dashboard['A1'].font = title_font
    ws_dashboard.merge_cells('A1:D1')
    
    ws_dashboard['A2'] = f"Tahun Analisis: {results.get('analysis_year', 'Semua Tahun')} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_dashboard['A2'].font = subtitle_font
    ws_dashboard.merge_cells('A2:D2')

    # Summary Section
    ws_dashboard['A4'] = "📈 Ringkasan Data"
    ws_dashboard['A4'].font = section_font
    ws_dashboard.merge_cells('A4:D4')

    # Headers
    headers = ["Metrik", "Jumlah Data", "Rata-rata Perubahan", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws_dashboard.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows
    row_data = [
        ("M ke M (Bulanan)", len(results.get('monthly_comparison', [])), _calc_avg_change(results.get('monthly_comparison', [])), "✓" if has_mm else "✗"),
        ("Q ke Q (Kuartal)", len(results.get('quarterly_comparison', [])), _calc_avg_change(results.get('quarterly_comparison', [])), "✓" if has_qq else "✗"),
        ("Y ke Y (Tahunan)", len(results.get('yearly_comparison', [])), _calc_avg_change(results.get('yearly_comparison', [])), "✓" if has_yy else "✗"),
        ("YTD (Akumulasi)", sum(len(y.get('monthly_ytd', [])) for y in results.get('ytd_comparison', [])), "-", "✓" if has_ytd else "✗"),
        ("C ke C (YoY)", len(results.get('current_to_current', [])), _calc_avg_change(results.get('current_to_current', [])), "✓" if has_cc else "✗"),
    ]

    for idx, (metric, count, avg_change, status) in enumerate(row_data, start=6):
        ws_dashboard.cell(row=idx, column=1, value=metric).border = thin_border
        ws_dashboard.cell(row=idx, column=2, value=count).border = thin_border
        ws_dashboard.cell(row=idx, column=2).alignment = center_align
        change_cell = ws_dashboard.cell(row=idx, column=3, value=avg_change)
        change_cell.border = thin_border
        change_cell.alignment = center_align
        status_cell = ws_dashboard.cell(row=idx, column=4, value=status)
        status_cell.border = thin_border
        status_cell.alignment = center_align

    # Column widths
    ws_dashboard.column_dimensions['A'].width = 22
    ws_dashboard.column_dimensions['B'].width = 14
    ws_dashboard.column_dimensions['C'].width = 20
    ws_dashboard.column_dimensions['D'].width = 10

    # ========== HELPER FUNCTION FOR DATA SHEETS ==========
    def create_data_sheet(sheet_name, headers, data_items, period_col_name):
        """Helper to create standardized data sheets with conditional formatting"""
        ws = wb.create_sheet(sheet_name)
        
        # Title
        ws['A1'] = f"{sheet_name} - {indicator}"
        ws['A1'].font = section_font
        ws.merge_cells(f'A1:E1')
        
        # Headers (row 3, leave row 2 empty for spacing)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        for idx, item in enumerate(data_items, start=4):
            ws.cell(row=idx, column=1, value=item['period']).border = thin_border
            ws.cell(row=idx, column=2, value=item['current_value']).border = thin_border
            ws.cell(row=idx, column=2).alignment = center_align
            ws.cell(row=idx, column=3, value=item['previous_value']).border = thin_border
            ws.cell(row=idx, column=3).alignment = center_align
            
            change_cell = ws.cell(row=idx, column=4, value=f"{item['change_percent']}%")
            change_cell.border = thin_border
            change_cell.alignment = center_align
            
            status_cell = ws.cell(row=idx, column=5, value=item['change_type'].capitalize())
            status_cell.border = thin_border
            status_cell.alignment = center_align

            # Conditional formatting (background color)
            if item['change_type'] == 'increase':
                change_cell.fill = green_fill
                change_cell.font = green_font
                status_cell.fill = green_fill
                status_cell.font = green_font
            elif item['change_type'] == 'decrease':
                change_cell.fill = red_fill
                change_cell.font = red_font
                status_cell.fill = red_fill
                status_cell.font = red_font

        # Column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 12
        
        return ws

    # ========== SHEET 2: M to M ==========
    if has_mm:
        create_data_sheet("M to M", ["Bulan", "Nilai", "Sebelumnya", "Perubahan", "Status"], 
                         results['monthly_comparison'], "Bulan")

    # ========== SHEET 3: Q to Q ==========
    if has_qq:
        create_data_sheet("Q ke Q", ["Kuartal", "Nilai", "Sebelumnya", "Perubahan", "Status"], 
                         results['quarterly_comparison'], "Kuartal")

    # ========== SHEET 4: Y to Y ==========
    if has_yy:
        create_data_sheet("Y ke Y", ["Tahun", "Nilai", "Sebelumnya", "Perubahan", "Status"], 
                         results['yearly_comparison'], "Tahun")

    # ========== SHEET 5: YTD ==========
    if has_ytd:
        ws_ytd = wb.create_sheet("YTD")
        ws_ytd['A1'] = f"YTD Analysis - {indicator}"
        ws_ytd['A1'].font = section_font
        ws_ytd.merge_cells('A1:D1')

        current_row = 3
        for year_data in results['ytd_comparison']:
            year = year_data.get('year', 'Unknown')
            
            # Year header
            ws_ytd.cell(row=current_row, column=1, value=f"Tahun {year}")
            ws_ytd.cell(row=current_row, column=1).font = Font(bold=True, size=11, color="366092")
            current_row += 1
            
            # Monthly YTD headers
            ytd_headers = ["Periode", "Nilai Bulanan", "YTD Akumulasi", "Status"]
            for col, header in enumerate(ytd_headers, 1):
                cell = ws_ytd.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1
            
            # Monthly data
            monthly_data = year_data.get('monthly_ytd', [])
            for item in monthly_data:
                ws_ytd.cell(row=current_row, column=1, value=item['period']).border = thin_border
                ws_ytd.cell(row=current_row, column=2, value=item.get('monthly_value', '-')).border = thin_border
                ws_ytd.cell(row=current_row, column=2).alignment = center_align
                ytd_cell = ws_ytd.cell(row=current_row, column=3, value=item['ytd_value'])
                ytd_cell.border = thin_border
                ytd_cell.alignment = center_align
                ytd_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                ws_ytd.cell(row=current_row, column=4, value="Akumulasi").border = thin_border
                current_row += 1
            
            current_row += 1  # Space between years

        # Column widths
        ws_ytd.column_dimensions['A'].width = 14
        ws_ytd.column_dimensions['B'].width = 16
        ws_ytd.column_dimensions['C'].width = 18
        ws_ytd.column_dimensions['D'].width = 14

    # ========== SHEET 6: C to C ==========
    if has_cc:
        ws_cc = wb.create_sheet("C ke C")
        ws_cc['A1'] = f"C ke C (Current vs Previous Year) - {indicator}"
        ws_cc['A1'].font = section_font
        ws_cc.merge_cells('A1:E1')
        
        headers = ["Periode", "Nilai Tahun Ini", "Nilai Tahun Lalu", "Perubahan (%)", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws_cc.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for idx, item in enumerate(results['current_to_current'], start=4):
            ws_cc.cell(row=idx, column=1, value=item['period']).border = thin_border
            ws_cc.cell(row=idx, column=2, value=item['current_value']).border = thin_border
            ws_cc.cell(row=idx, column=2).alignment = center_align
            ws_cc.cell(row=idx, column=3, value=item['previous_year_value']).border = thin_border
            ws_cc.cell(row=idx, column=3).alignment = center_align
            
            change_cell = ws_cc.cell(row=idx, column=4, value=f"{item['change_percent']}%")
            change_cell.border = thin_border
            change_cell.alignment = center_align
            
            status_cell = ws_cc.cell(row=idx, column=5, value=item['change_type'].capitalize())
            status_cell.border = thin_border
            status_cell.alignment = center_align

            # Conditional formatting
            if item['change_type'] == 'increase':
                change_cell.fill = green_fill
                change_cell.font = green_font
                status_cell.fill = green_fill
                status_cell.font = green_font
            elif item['change_type'] == 'decrease':
                change_cell.fill = red_fill
                change_cell.font = red_font
                status_cell.fill = red_fill
                status_cell.font = red_font

        # Column widths
        ws_cc.column_dimensions['A'].width = 14
        ws_cc.column_dimensions['B'].width = 18
        ws_cc.column_dimensions['C'].width = 18
        ws_cc.column_dimensions['D'].width = 16
        ws_cc.column_dimensions['E'].width = 12

    # ========== SHEET 7: METADATA ==========
    ws_meta = wb.create_sheet("Metadata")
    ws_meta['A1'] = "📋 Informasi Analisis"
    ws_meta['A1'].font = title_font
    ws_meta.merge_cells('A1:C1')

    metadata_items = [
        ("Indikator", indicator),
        ("Tahun Analisis", results.get('analysis_year', 'Semua Tahun')),
        ("Tanggal Export", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ("Jumlah Periode M-M", str(len(results.get('monthly_comparison', [])))),
        ("Jumlah Periode Q-Q", str(len(results.get('quarterly_comparison', [])))),
        ("Jumlah Periode Y-Y", str(len(results.get('yearly_comparison', [])))),
        ("Jumlah Data C-C", str(len(results.get('current_to_current', [])))),
        ("", ""),
        ("Catatan", "Data ini dihasilkan oleh Sistem Data BPS"),
        ("Disclaimer", "Data untuk analisis internal. Validasi ke sumber resmi BPS sebelum dipublikasikan."),
    ]

    for idx, (label, value) in enumerate(metadata_items, start=3):
        label_cell = ws_meta.cell(row=idx, column=1, value=label)
        label_cell.font = Font(bold=True if label else False)
        value_cell = ws_meta.cell(row=idx, column=2, value=value)
        if label == "Catatan" or label == "Disclaimer":
            value_cell.font = Font(italic=True, size=9)
            ws_meta.merge_cells(f'B{idx}:C{idx}')

    ws_meta.column_dimensions['A'].width = 20
    ws_meta.column_dimensions['B'].width = 50
    ws_meta.column_dimensions['C'].width = 20

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"BPS_Analysis_{indicator}_{timestamp}.xlsx"

    return Response(
        output.getvalue(),
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )


def _calc_avg_change(comparison_list):
    """Helper function to calculate average change percentage"""
    if not comparison_list:
        return "N/A"
    try:
        avg = sum(item.get('change_percent', 0) for item in comparison_list) / len(comparison_list)
        return f"{avg:+.2f}%"
    except:
        return "N/A"


if __name__ == "__main__":
    app.run(debug=True, port=int(os.environ.get("FLASK_RUN_PORT", 5000)))
