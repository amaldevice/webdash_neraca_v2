"""
Test Edge Cases - BPS Data Management System
Testing untuk empty data, large files, encoding issues, dan network timeouts
"""

import pytest
import os
import tempfile
import time
from io import BytesIO
from unittest.mock import patch, MagicMock
import pandas as pd
import threading


class TestEdgeCases:
    """Test class untuk edge cases"""

    def test_empty_excel_file(self, test_client):
        """Test upload file Excel kosong"""
        # Create empty Excel file
        df = pd.DataFrame()
        excel_buffer = BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)

        with test_client as client:
            data = {
                'uploader': 'TestUser',
                'version': 'v1.0',
                'data_type': 'flow',
                'time_period': 'monthly'
            }
            file_data = (excel_buffer, 'empty.xlsx')

            response = client.post('/upload',
                                 data=data,
                                 content_type='multipart/form-data',
                                 data_file=file_data)

            assert response.status_code in [200, 302]
            with client.session_transaction() as sess:
                flashed_messages = sess.get('_flashes', [])
                response_text = response.get_data(as_text=True).lower()
                if flashed_messages:
                    error_found = any("tidak berisi data" in msg[1] or "empty" in msg[1].lower()
                                    for msg in flashed_messages)
                    assert error_found
                else:
                    assert "tidak berisi data" in response_text or "empty" in response_text or "kosong" in response_text or "error" in response_text

    def test_excel_file_with_only_headers(self, test_client):
        """Test upload file Excel yang hanya berisi header"""
        df = pd.DataFrame(columns=['Indicator', '2024-01', '2024-02'])
        excel_buffer = BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)

        with test_client as client:
            data = {
                'uploader': 'TestUser',
                'version': 'v1.0',
                'data_type': 'flow',
                'time_period': 'monthly'
            }
            file_data = (excel_buffer, 'headers_only.xlsx')

            response = client.post('/upload',
                                 data=data,
                                 content_type='multipart/form-data',
                                 data_file=file_data)

            assert response.status_code in [200, 302]
            with client.session_transaction() as sess:
                flashed_messages = sess.get('_flashes', [])
                if not flashed_messages:
                    response_text = response.get_data(as_text=True).lower()
                    assert "error" in response_text or "header" in response_text

    def test_excel_file_with_empty_rows(self, test_client):
        """Test upload file Excel dengan baris kosong"""
        df = pd.DataFrame({
            'Indicator': ['GDP', '', 'Inflation', None],
            '2024-01': [100, None, 200, '']
        })
        excel_buffer = BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)

        with test_client as client:
            data = {
                'uploader': 'TestUser',
                'version': 'v1.0',
                'data_type': 'flow',
                'time_period': 'monthly'
            }
            file_data = (excel_buffer, 'empty_rows.xlsx')

            response = client.post('/upload',
                                 data=data,
                                 content_type='multipart/form-data',
                                 data_file=file_data)

            # Should handle gracefully - only valid rows should be processed
            assert response.status_code in [200, 302]

    def test_very_large_excel_file(self, large_excel_file, test_client):
        """Test upload file Excel yang sangat besar (melebihi batas 16MB)"""
        with test_client as client:
            data = {
                'uploader': 'TestUser',
                'version': 'v1.0',
                'data_type': 'flow',
                'time_period': 'monthly'
            }
            # Read the large file
            with open(large_excel_file, 'rb') as f:
                file_content = f.read()

            file_data = (BytesIO(file_content), 'large.xlsx')

            response = client.post('/upload',
                                 data=data,
                                 content_type='multipart/form-data',
                                 data_file=file_data)

            # Should be rejected due to size limit
            assert response.status_code == 413 or response.status_code == 302

            if response.status_code == 302:
                with client.session_transaction() as sess:
                    flashed_messages = sess.get('_flashes', [])
                    if not flashed_messages:
                        response_text = response.get_data(as_text=True).lower()
                        assert "error" in response_text or "terlalu besar" in response_text or "413" in response_text

    def test_excel_file_with_special_characters(self, test_client):
        """Test upload file Excel dengan karakter spesial"""
        special_indicators = [
            "GDP©®™",
            "Inflationñáéíóú",
            "人口统计",  # Chinese
            "статистика",  # Russian
            "στατιστική",  # Greek
            "GDP ± ∞ ÷ ×"
        ]

        for indicator in special_indicators:
            with test_client as client:
                df = pd.DataFrame({indicator: [100], '2024-01': [100]})
                excel_buffer = BytesIO()
                df.to_excel(excel_buffer, index=False)
                excel_buffer.seek(0)

                data = {
                    'uploader': 'TestUser',
                    'version': 'v1.0',
                    'data_type': 'flow',
                    'time_period': 'monthly'
                }
                file_data = (excel_buffer, 'special_chars.xlsx')

                response = client.post('/upload',
                                     data=data,
                                     content_type='multipart/form-data',
                                     data_file=file_data)

                # Should handle special characters
                assert response.status_code in [200, 302]

    def test_excel_file_with_unicode_encoding(self, test_client):
        """Test upload file Excel dengan encoding Unicode"""
        unicode_indicators = [
            "GDP_ñáéíóú_中文_русский",
            "Inflation_🚀⭐📊💹",
            "人口_🌍🌎🌏"
        ]

        for indicator in unicode_indicators:
            with test_client as client:
                df = pd.DataFrame({indicator: [100], '2024-01': [100]})
                excel_buffer = BytesIO()
                df.to_excel(excel_buffer, index=False)
                excel_buffer.seek(0)

                data = {
                    'uploader': 'TestUser',
                    'version': 'v1.0',
                    'data_type': 'flow',
                    'time_period': 'monthly'
                }
                file_data = (excel_buffer, 'unicode.xlsx')

                response = client.post('/upload',
                                     data=data,
                                     content_type='multipart/form-data',
                                     data_file=file_data)

                assert response.status_code in [200, 302]

    def test_excel_file_corrupted_during_upload(self, test_client):
        """Test simulasi file Excel yang korup saat upload"""
        # Create a valid Excel file then corrupt it
        df = pd.DataFrame({'Indicator': ['GDP'], '2024-01': [100]})
        excel_buffer = BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)

        # Corrupt the file by modifying bytes
        corrupted_content = excel_buffer.getvalue()
        corrupted_content = corrupted_content[:100] + b'\x00\x01\x02' + corrupted_content[103:]

        with test_client as client:
            data = {
                'uploader': 'TestUser',
                'version': 'v1.0',
                'data_type': 'flow',
                'time_period': 'monthly'
            }
            file_data = (BytesIO(corrupted_content), 'corrupted.xlsx')

            response = client.post('/upload',
                                 data=data,
                                 content_type='multipart/form-data',
                                 data_file=file_data)

            # Should handle corruption gracefully
            assert response.status_code in [200, 302, 500]

    def test_network_timeout_simulation(self, test_client):
        """Test simulasi network timeout"""
        import time

        with patch('excel_parser.parse_excel') as mock_parse:
            def slow_parse(*args, **kwargs):
                time.sleep(2)  # Simulate network delay
                return [{'uploader_name': 'Test', 'version': 'v1.0', 'indicator_name': 'GDP', 'value': 100}]
            mock_parse.side_effect = slow_parse

            with test_client as client:
                df = pd.DataFrame({'Indicator': ['GDP'], '2024-01': [100]})
                excel_buffer = BytesIO()
                df.to_excel(excel_buffer, index=False)
                excel_buffer.seek(0)

                data = {
                    'uploader': 'TestUser',
                    'version': 'v1.0',
                    'data_type': 'flow',
                    'time_period': 'monthly'
                }
                file_data = (excel_buffer, 'test.xlsx')

                start_time = time.time()
                response = client.post('/upload',
                                     data=data,
                                     content_type='multipart/form-data',
                                     data_file=file_data)
                end_time = time.time()

                # Should complete within reasonable time
                assert end_time - start_time < 10  # Less than 10 seconds
                assert response.status_code in [200, 302]

    def test_database_connection_timeout(self, test_client):
        """Test simulasi database connection timeout"""
        with patch("infrastructure.db.get_session") as mock_get_session:
            def timeout_conn():
                time.sleep(1)  # Simulate connection delay
                raise Exception("Connection timeout")

            mock_get_session.side_effect = timeout_conn

            with test_client as client:
                df = pd.DataFrame({'Indicator': ['GDP'], '2024-01': [100]})
                excel_buffer = BytesIO()
                df.to_excel(excel_buffer, index=False)
                excel_buffer.seek(0)

                data = {
                    'uploader': 'TestUser',
                    'version': 'v1.0',
                    'data_type': 'flow',
                    'time_period': 'monthly'
                }
                file_data = (excel_buffer, 'test.xlsx')

                response = client.post('/upload',
                                     data=data,
                                     content_type='multipart/form-data',
                                     data_file=file_data)

                # Should handle timeout gracefully
                assert response.status_code in [200, 302, 500]

    def test_file_system_full_disk(self, test_client):
        """Test simulasi disk penuh saat menyimpan file"""
        with test_client as client:
            df = pd.DataFrame({'Indicator': ['GDP'], '2024-01': [100]})
            excel_buffer = BytesIO()
            df.to_excel(excel_buffer, index=False)
            excel_buffer.seek(0)

            data = {
                'uploader': 'TestUser',
                'version': 'v1.0',
                'data_type': 'flow',
                'time_period': 'monthly'
            }
            file_data = (excel_buffer, 'test.xlsx')

            with patch('services.upload_flow.open', side_effect=OSError("No space left on device")):
                response = client.post('/upload',
                                     data=data,
                                     content_type='multipart/form-data',
                                     data_file=file_data)

                # Should handle disk full error gracefully
                assert response.status_code in [200, 302, 500]

    def test_concurrent_file_access(self, test_client):
        """Test akses file bersamaan dari multiple threads"""
        results = []
        errors = []

        def upload_worker(worker_id):
            try:
                with test_client.application.test_client() as client:
                    df = pd.DataFrame({'Indicator': [f'GDP_{worker_id}'], '2024-01': [100 + worker_id]})
                    excel_buffer = BytesIO()
                    df.to_excel(excel_buffer, index=False)
                    excel_buffer.seek(0)

                    data = {
                        'uploader': f'TestUser_{worker_id}',
                        'version': 'v1.0',
                        'data_type': 'flow',
                        'time_period': 'monthly'
                    }
                    file_data = (excel_buffer, f'test_{worker_id}.xlsx')

                    response = client.post('/upload',
                                         data=data,
                                         content_type='multipart/form-data',
                                         data_file=file_data)

                    results.append((worker_id, response.status_code))
            except Exception as e:
                errors.append((worker_id, str(e)))

        # Start multiple threads
        threads = []
        for i in range(5):
            thread = threading.Thread(target=upload_worker, args=(i,))
            threads.append(thread)
            thread.start()

        # Wait for all threads to complete
        for thread in threads:
            thread.join()

        # All uploads should succeed
        assert len(results) == 5
        for worker_id, status_code in results:
            assert status_code == 302, f"Worker {worker_id} failed with status {status_code}"

        assert len(errors) == 0, f"Errors occurred: {errors}"

    def test_extremely_nested_excel_structure(self, test_client):
        """Test file Excel dengan struktur sangat kompleks"""
        # Create Excel with many sheets, complex formulas, etc.
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Sheet 1: Normal data
            df1 = pd.DataFrame({'Indicator': ['GDP'], '2024-01': [100]})
            df1.to_excel(writer, sheet_name='Sheet1', index=False)

            # Sheet 2: Empty
            df2 = pd.DataFrame()
            df2.to_excel(writer, sheet_name='Sheet2', index=False)

            # Sheet 3: Complex data
            df3 = pd.DataFrame({
                'Indicator': ['GDP', 'Inflation', 'Population'],
                'Q1': [100, 2.5, 1000000],
                'Q2': [105, 2.8, 1005000],
                'Q3': [98, 2.3, 1010000],
                'Q4': [110, 3.0, 1015000]
            })
            df3.to_excel(writer, sheet_name='Sheet3', index=False)

        buffer.seek(0)

        with test_client as client:
            data = {
                'uploader': 'TestUser',
                'version': 'v1.0',
                'data_type': 'flow',
                'time_period': 'quarterly'
            }
            file_data = (buffer, 'complex.xlsx')

            response = client.post('/upload',
                                 data=data,
                                 content_type='multipart/form-data',
                                 data_file=file_data)

            # Should handle complex Excel files
            assert response.status_code in [200, 302]

    def test_excel_with_merged_cells(self, test_client):
        """Test file Excel dengan merged cells"""
        # Create Excel with merged cells using openpyxl directly
        from openpyxl import Workbook
        from openpyxl.styles import Alignment

        wb = Workbook()
        ws = wb.active

        # Add data with merged cells
        ws['A1'] = 'Indicator'
        ws['B1'] = '2024'
        ws.merge_cells('B1:D1')  # Merge B1, C1, D1

        ws['A2'] = 'GDP'
        ws['B2'] = 100
        ws['C2'] = 105
        ws['D2'] = 110

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        with test_client as client:
            data = {
                'uploader': 'TestUser',
                'version': 'v1.0',
                'data_type': 'flow',
                'time_period': 'monthly'
            }
            file_data = (buffer, 'merged.xlsx')

            response = client.post('/upload',
                                 data=data,
                                 content_type='multipart/form-data',
                                 data_file=file_data)

            # Should handle merged cells
            assert response.status_code in [200, 302]

    def test_excel_with_formulas(self, test_client):
        """Test file Excel dengan formulas"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Add data with formulas
        ws['A1'] = 'Indicator'
        ws['B1'] = 'Value1'
        ws['C1'] = 'Value2'
        ws['D1'] = 'Sum'

        ws['A2'] = 'GDP'
        ws['B2'] = 100
        ws['C2'] = 50
        ws['D2'] = '=B2+C2'  # Formula

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        with test_client as client:
            data = {
                'uploader': 'TestUser',
                'version': 'v1.0',
                'data_type': 'flow',
                'time_period': 'monthly'
            }
            file_data = (buffer, 'formulas.xlsx')

            response = client.post('/upload',
                                 data=data,
                                 content_type='multipart/form-data',
                                 data_file=file_data)

            # Should handle formulas (pandas should get computed values)
            assert response.status_code in [200, 302]

    def test_empty_database_queries(self, test_client):
        """Test query pada database kosong"""
        # Test various endpoints with empty database
        with test_client as client:
            # Landing page
            response = client.get('/')
            assert response.status_code == 200

            # Preview data
            response = client.get('/preview-data')
            assert response.status_code == 200

            # Export
            response = client.get('/export')
            assert response.status_code == 200

    def test_pagination_edge_cases(self, populated_db):
        """Test pagination dengan edge cases"""
        with populated_db as client:
            # Test page 0
            response = client.get('/preview-data?page=0&limit=10')
            assert response.status_code == 200

            # Test negative page
            response = client.get('/preview-data?page=-1&limit=10')
            assert response.status_code == 200

            # Test very high page number
            response = client.get('/preview-data?page=99999&limit=10')
            assert response.status_code == 200

            # Test limit 0
            response = client.get('/preview-data?page=1&limit=0')
            assert response.status_code == 200

            # Test negative limit
            response = client.get('/preview-data?page=1&limit=-5')
            assert response.status_code == 200

    def test_export_with_no_data(self, test_client):
        """Test export ketika tidak ada data"""
        with test_client as client:
            response = client.get('/export?format=csv')
            assert response.status_code == 200
            assert 'Content-Disposition' in response.headers

            response = client.get('/export?format=excel')
            assert response.status_code == 200
            assert 'Content-Disposition' in response.headers

    def test_chart_generation_with_no_data(self, test_client):
        """Test generate chart ketika tidak ada data"""
        with test_client as client:
            data = {'indicator_filter': 'NonExistent', 'time_range': 'all'}
            response = client.post('/generate-plot', data=data)
            assert response.status_code == 200

            data = {'indicator': 'NonExistent', 'year': '2024'}
            response = client.post('/generate-period-analysis', data=data)
            assert response.status_code == 200